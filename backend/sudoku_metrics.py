from __future__ import annotations
import copy
import csv
import json
import time
import tracemalloc
from pathlib import Path
from typing import Any, Callable
from sudoku_board import puzzle_a_tablero, verificar_solucion_valida
from sudoku_solvers import ALGORITMOS, SolverMetrics

def medir_tiempo_ejecucion(funcion: Callable[[], Any]) -> tuple[Any, float]:
    inicio = time.perf_counter()
    resultado = funcion()
    tiempo_ms = (time.perf_counter() - inicio) * 1000
    return (resultado, tiempo_ms)

def medir_consumo_memoria(funcion: Callable[[], Any]) -> tuple[Any, float]:
    tracemalloc.start()
    try:
        resultado = funcion()
        _, pico = tracemalloc.get_traced_memory()
    finally:
        tracemalloc.stop()
    return (resultado, round(pico / 1024, 3))

def resolver_con_metricas(puzzle: list[list[Any]], size: int, algoritmo: str) -> dict[str, Any]:
    if algoritmo not in ALGORITMOS:
        raise ValueError(f'Algoritmo desconocido: {algoritmo}')
    inicial = puzzle_a_tablero(puzzle, size)
    tablero = copy.deepcopy(inicial)
    metrics = SolverMetrics()
    info = ALGORITMOS[algoritmo]
    resolver = info['fn']
    tracemalloc.start()
    inicio = time.perf_counter()
    exito = resolver(tablero, size, metrics)
    tiempo_ms = (time.perf_counter() - inicio) * 1000
    _, pico = tracemalloc.get_traced_memory()
    tracemalloc.stop()
    solucion_valida = exito and verificar_solucion_valida(tablero, size)
    return {'success': solucion_valida, 'solution': tablero, 'metrics': {'execution_time_ms': round(tiempo_ms, 3), 'nodes_explored': metrics.nodes, 'backtracks': metrics.backtracks, 'combinations_tried': metrics.combinations_tried, 'memory_peak_kb': round(pico / 1024, 3), 'algorithm': info['name'], 'algorithm_key': algoritmo}}

def comparar_rendimiento_algoritmos(puzzle: list[list[Any]], size: int, algoritmos: list[str] | None=None) -> dict[str, Any]:
    claves = algoritmos or ['backtracking', 'mrv', 'brute_force']
    resultados: list[dict[str, Any]] = []
    for clave in claves:
        if clave not in ALGORITMOS:
            continue
        if clave == 'brute_force' and size > 9:
            resultados.append({'algorithm_key': clave, 'algorithm': ALGORITMOS[clave]['name'], 'success': False, 'skipped': True, 'reason': 'Fuerza Bruta solo disponible para 9×9.'})
            continue
        resultado = resolver_con_metricas(puzzle, size, clave)
        metricas = resultado['metrics']
        resultados.append({'algorithm_key': clave, 'algorithm': metricas['algorithm'], 'success': resultado['success'], 'execution_time_ms': metricas['execution_time_ms'], 'nodes_explored': metricas['nodes_explored'], 'backtracks': metricas['backtracks'], 'combinations_tried': metricas['combinations_tried'], 'memory_peak_kb': metricas['memory_peak_kb']})
    tiempos = [r['execution_time_ms'] for r in resultados if r.get('success')]
    mas_rapido = min(tiempos) if tiempos else None
    for fila in resultados:
        if fila.get('success') and mas_rapido:
            fila['speedup_vs_slowest'] = round(max(tiempos) / fila['execution_time_ms'], 2)
    return {'size': size, 'comparison': resultados, 'fastest_time_ms': mas_rapido}

def calcular_porcentaje_resueltos(puzzles: list[dict[str, Any]], algoritmo: str='backtracking') -> dict[str, Any]:
    total = len(puzzles)
    resueltos = 0
    detalle: list[dict[str, Any]] = []
    for indice, item in enumerate(puzzles):
        size = item['size']
        puzzle = item['puzzle']
        try:
            resultado = resolver_con_metricas(puzzle, size, algoritmo)
            exito = resultado['success']
        except ValueError:
            exito = False
        if exito:
            resueltos += 1
        detalle.append({'index': indice, 'success': exito})
    porcentaje = round(resueltos / total * 100, 2) if total else 0.0
    return {'algorithm': algoritmo, 'total_boards': total, 'solved_correctly': resueltos, 'success_rate_percent': porcentaje, 'details': detalle}

def exportar_metricas(metricas: dict[str, Any], ruta_salida: str | Path, formato: str='json') -> str:
    destino = Path(ruta_salida)
    destino.parent.mkdir(parents=True, exist_ok=True)
    if formato == 'json':
        if destino.suffix.lower() != '.json':
            destino = destino.with_suffix('.json')
        destino.write_text(json.dumps(metricas, indent=2, ensure_ascii=False), encoding='utf-8')
        return str(destino.resolve())
    if formato == 'csv':
        if destino.suffix.lower() != '.csv':
            destino = destino.with_suffix('.csv')
        filas = metricas.get('comparison') or [metricas.get('metrics', metricas)]
        if not filas:
            raise ValueError('No hay métricas para exportar.')
        campos = sorted({clave for fila in filas for clave in fila.keys()})
        with destino.open('w', newline='', encoding='utf-8') as archivo:
            escritor = csv.DictWriter(archivo, fieldnames=campos)
            escritor.writeheader()
            for fila in filas:
                escritor.writerow(fila)
        return str(destino.resolve())
    if formato == 'xlsx':
        return _exportar_metricas_excel(metricas, destino)
    raise ValueError("Formato no soportado. Use 'json', 'csv' o 'xlsx'.")

def _exportar_metricas_excel(metricas: dict[str, Any], destino: Path) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    if destino.suffix.lower() != '.xlsx':
        destino = destino.with_suffix('.xlsx')
    wb = Workbook()
    hoja_resumen = wb.active
    hoja_resumen.title = 'Resumen'
    encabezado_fill = PatternFill('solid', fgColor='283593')
    encabezado_font = Font(bold=True, color='FFFFFF')
    filas = metricas.get('comparison') or [metricas.get('metrics', metricas)]
    if not filas:
        raise ValueError('No hay métricas para exportar.')
    campos = ['algorithm', 'algorithm_key', 'success', 'execution_time_ms', 'nodes_explored', 'backtracks', 'combinations_tried', 'memory_peak_kb', 'speedup_vs_slowest']
    campos_presentes = [c for c in campos if any((c in f for f in filas))]
    hoja_resumen.append(campos_presentes)
    for celda in hoja_resumen[1]:
        celda.fill = encabezado_fill
        celda.font = encabezado_font
    for fila in filas:
        hoja_resumen.append([fila.get(c, '') for c in campos_presentes])
    if 'success_rate_percent' in metricas:
        hoja_benchmark = wb.create_sheet('Benchmark')
        hoja_benchmark.append(['Métrica', 'Valor'])
        hoja_benchmark.append(['Algoritmo', metricas.get('algorithm', '')])
        hoja_benchmark.append(['Total tableros', metricas.get('total_boards', '')])
        hoja_benchmark.append(['Resueltos correctamente', metricas.get('solved_correctly', '')])
        hoja_benchmark.append(['Porcentaje de éxito (%)', metricas.get('success_rate_percent', '')])
    if metricas.get('metrics') and (not metricas.get('comparison')):
        m = metricas['metrics']
        hoja_resumen['J1'] = 'Ejecución individual'
        hoja_resumen['J2'] = m.get('algorithm', '')
    for columna in hoja_resumen.columns:
        max_len = max((len(str(c.value or '')) for c in columna))
        hoja_resumen.column_dimensions[columna[0].column_letter].width = min(max_len + 2, 40)
    wb.save(destino)
    return str(destino.resolve())
