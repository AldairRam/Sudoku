"""
Carga y consulta de datasets de Sudoku.

Requisitos cubiertos:
  1. Cargar los datasets de Sudoku (desde Excel).
  2. Obtener la lista de datasets disponibles.
  3. Obtener un tablero de Sudoku por su ID.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from sudoku_constants import SYMBOLS
from sudoku_board import sanitizar_puzzle

DATA_DIR = Path(__file__).parent / "data"
DEFAULT_EXCEL_PATH = DATA_DIR / "SudokuDatos.xlsx"
FALLBACK_JSON_PATH = DATA_DIR / "datasets.json"
EXCEL_SHEET = "Datos_Sudoku"

_DATASETS: dict[str, dict[str, Any]] | None = None


def _parsear_matriz_string(matriz: str, size: int) -> list[int]:
    """
    Convierte la cadena del Excel en una lista plana de enteros.
    '.' representa celda vacía (0). Para 12×12 y 16×16, A–G equivalen a 10–16.
    """
    letras = {chr(ord("A") + i): 10 + i for i in range(7)}
    valores: list[int] = []
    total = size * size

    for caracter in matriz:
        if len(valores) >= total:
            break
        if caracter == ".":
            valores.append(0)
        elif caracter.isdigit():
            valores.append(int(caracter))
        elif caracter.upper() in letras:
            valores.append(letras[caracter.upper()])

    if len(valores) != total:
        raise ValueError(
            f"La matriz {matriz[:24]}… no produce {total} celdas (obtenidas: {len(valores)})."
        )
    return valores


def matriz_string_a_puzzle(matriz: str, size: int) -> list[list[Any]]:
    """Convierte Matriz_String del Excel en matriz bidimensional para la API."""
    plana = _parsear_matriz_string(matriz, size)
    puzzle: list[list[Any]] = []

    for fila in range(size):
        fila_datos: list[Any] = []
        for columna in range(size):
            valor = plana[fila * size + columna]
            if valor == 0:
                fila_datos.append(0)
            elif size >= 12 and valor in SYMBOLS:
                fila_datos.append(SYMBOLS[valor])
            else:
                fila_datos.append(valor)
        puzzle.append(fila_datos)

    return puzzle


def _cargar_desde_excel(ruta: Path) -> dict[str, dict[str, Any]]:
    from openpyxl import load_workbook

    if not ruta.exists():
        raise FileNotFoundError(f"No se encontró el archivo Excel: {ruta}")

    libro = load_workbook(ruta, data_only=True, read_only=True)
    if EXCEL_SHEET not in libro.sheetnames:
        raise ValueError(f"No existe la hoja '{EXCEL_SHEET}' en {ruta.name}.")

    hoja = libro[EXCEL_SHEET]
    datasets: dict[str, dict[str, Any]] = {}

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if not fila or not fila[0]:
            continue

        dataset_id = str(fila[0]).strip()
        size = int(fila[1])
        dificultad = str(fila[2]).strip().capitalize()
        matriz = str(fila[3]).strip()

        puzzle = matriz_string_a_puzzle(matriz, size)
        puzzle, removidas = sanitizar_puzzle(puzzle, size)

        datasets[dataset_id] = {
            "id": dataset_id,
            "name": f"{size}×{size} — {dificultad} ({dataset_id})",
            "size": size,
            "difficulty": dificultad,
            "puzzle": puzzle,
            "source": "excel",
            "sanitized_clues_removed": removidas,
        }

    libro.close()
    return datasets


def _cargar_desde_json(ruta: Path) -> dict[str, dict[str, Any]]:
    with ruta.open(encoding="utf-8") as archivo:
        datos = json.load(archivo)
    if not isinstance(datos, dict):
        raise ValueError("El archivo JSON de datasets debe ser un objeto.")
    return datos


def cargar_datasets(ruta: str | Path | None = None) -> dict[str, dict[str, Any]]:
    """
    Carga datasets desde Excel (prioridad) o JSON de respaldo.
    Por defecto usa backend/data/SudokuDatos.xlsx.
    """
    global _DATASETS

    path = Path(ruta) if ruta else DEFAULT_EXCEL_PATH

    if path.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
        _DATASETS = _cargar_desde_excel(path)
    elif path.suffix.lower() == ".json":
        _DATASETS = _cargar_desde_json(path)
    elif DEFAULT_EXCEL_PATH.exists():
        _DATASETS = _cargar_desde_excel(DEFAULT_EXCEL_PATH)
    elif FALLBACK_JSON_PATH.exists():
        _DATASETS = _cargar_desde_json(FALLBACK_JSON_PATH)
    else:
        raise FileNotFoundError(
            f"No se encontró {DEFAULT_EXCEL_PATH} ni {FALLBACK_JSON_PATH}."
        )

    return _DATASETS


def obtener_datasets() -> dict[str, dict[str, Any]]:
    """Devuelve los datasets en memoria, cargándolos si aún no están disponibles."""
    global _DATASETS
    if _DATASETS is None:
        cargar_datasets()
    return _DATASETS


def listar_datasets_disponibles(size: int | None = None) -> list[dict[str, Any]]:
    """
    Devuelve metadatos de los datasets (id, name, size, difficulty).
    Si size está definido, filtra por ese tamaño de tablero.
    """
    datasets = obtener_datasets()
    resultado = [
        {
            "id": ds["id"],
            "name": ds["name"],
            "size": ds["size"],
            "difficulty": ds.get("difficulty"),
        }
        for ds in datasets.values()
    ]
    if size is not None:
        resultado = [ds for ds in resultado if ds["size"] == size]
    return sorted(resultado, key=lambda ds: ds["id"])


def obtener_tablero_por_id(dataset_id: str) -> dict[str, Any]:
    """Obtiene un dataset completo (incluye puzzle) a partir de su identificador."""
    datasets = obtener_datasets()
    if dataset_id not in datasets:
        raise KeyError(f"Dataset no encontrado: {dataset_id}")
    return datasets[dataset_id]
